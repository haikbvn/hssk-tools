"""Read an Excel workbook into raw ``{header: value}`` rows using openpyxl."""

from __future__ import annotations

from collections import Counter
from collections.abc import Callable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..errors import ConfigError
from ..events import MessageCode, Msg, render_en
from ..mapping import MappingConfig


def _clean_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def read_rows(
    path: str | Path,
    mapping: MappingConfig,
    *,
    on_warning: Callable[[Msg], None] | None = None,
) -> list[tuple[int, dict[str, Any]]]:
    """Return ``[(excel_row_number, {header: value})]`` for every non-empty data row.

    If ``on_warning`` is given, non-fatal header diagnostics (e.g. Excel columns not present in the
    mapping, which are silently ignored) are reported through it. The return value is unaffected.
    """
    p = Path(path)
    if not p.exists():
        raise ConfigError(f"Excel file not found: {p}")
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        ws = wb[mapping.sheet] if mapping.sheet else wb[wb.sheetnames[0]]
    except KeyError as exc:
        raise ConfigError(f"Sheet {mapping.sheet!r} not found. Available: {wb.sheetnames}") from exc

    headers: list[str] = []
    rows: list[tuple[int, dict[str, Any]]] = []
    try:
        for r_idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx < mapping.header_row:
                continue
            if r_idx == mapping.header_row:
                headers = [_clean_header(v) for v in values]
                continue
            if all(_is_blank(v) for v in values):
                continue
            raw = {h: v for h, v in zip(headers, values, strict=False) if h}
            rows.append((r_idx, raw))
    finally:
        wb.close()

    if not headers:
        raise ConfigError(f"No header row found at row {mapping.header_row} in {p}")
    _check_columns(headers, mapping, p)
    if on_warning is not None:
        for warning in check_headers(headers, mapping):
            on_warning(warning)
    return rows


def check_headers(headers: list[str], mapping: MappingConfig) -> list[Msg]:
    """Non-fatal header diagnostics: Excel columns not in the mapping that will be ignored.

    Returns at most one combined message so spreadsheets with many decorative columns stay quiet
    in the log/table.
    """
    seen: set[str] = set()
    extra: list[str] = []
    for h in headers:
        if h and h not in mapping.columns and h not in seen:
            seen.add(h)
            extra.append(h)
    if not extra:
        return []
    cols = ", ".join(repr(h) for h in extra)
    return [Msg(MessageCode.LOG_UNMAPPED_COLUMNS, {"n": len(extra), "cols": cols})]


def _check_columns(headers: list[str], mapping: MappingConfig, path: Path) -> None:
    present = set(headers)
    missing = [c for c in mapping.columns if c not in present]
    if missing:
        msg = Msg(
            MessageCode.FILE_MISSING_COLUMNS,
            {"name": path.name, "missing": missing, "headers": headers},
        )
        raise ConfigError(render_en(msg), msg=msg)
    dups = [h for h, n in Counter(h for h in headers if h in mapping.columns).items() if n > 1]
    if dups:
        msg = Msg(MessageCode.FILE_DUPLICATE_COLUMNS, {"name": path.name, "dups": dups})
        raise ConfigError(render_en(msg), msg=msg)
