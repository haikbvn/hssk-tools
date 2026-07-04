"""Generate a fill-in Excel template whose headers exactly match a mapping."""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from ..mapping import ColumnSpec, MappingConfig
from .coerce import _RANGES  # single source of truth for vital soft-ranges

# ── Shared fill + text-color constants (reused by guide-sheet legend) ────────
# Light fills with dark text — high legibility, familiar Excel Good/Bad look.
_FILL_ID = PatternFill("solid", fgColor="FFE699")  # light amber  — identifier
_FILL_REQ = PatternFill("solid", fgColor="FFC7CE")  # light red    — required
_FILL_OPT = PatternFill("solid", fgColor="DDEBF7")  # light blue   — optional

_COLOR_ID = "7F6000"  # dark amber text for identifier header
_COLOR_REQ = "9C0006"  # dark red   text for required headers
_COLOR_OPT = "1F4E78"  # dark navy  text for optional headers

# ── Border constants ──────────────────────────────────────────────────────────
_SIDE_THIN = Side(style="thin", color="BFBFBF")
_SIDE_MED = Side(style="medium", color="808080")
_BORDER_CELL = Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_THIN)
# Header cells get a medium bottom border to visually separate from data rows.
_BORDER_HEADER = Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=_SIDE_MED)

# Rows 2..(1+_DATA_ROWS) are the editable data block (unlocked + validated).
_DATA_ROWS = 1000

# Per-column cell number formats for the data region (guard against Excel mangling input).
_FMT_TEXT = "@"  # keep as text: preserves leading zeros (CCCD) and avoids scientific notation
_FMT_DATETIME = "dd/mm/yyyy hh:mm"
_FMT_INT = "0"


def _number_format(spec: ColumnSpec, *, is_identifier: bool) -> str | None:
    """Data-region number format for a column; None keeps Excel's General format.

    float/str_num deliberately stay General: those columns carry ``decimal`` data-validation rules,
    and a Text format would store every entry as a string and trip the DV popup. Coercion already
    accepts both numeric cells and VN comma-decimal strings, so nothing is lost.
    """
    if is_identifier or spec.type in ("str", "list"):
        return _FMT_TEXT
    if spec.type == "datetime":
        return _FMT_DATETIME
    if spec.type == "int":
        return _FMT_INT
    return None


# API targets that are integer codes (full valid sets unknown → whole-number STOP only)
_CODE_INT_TARGETS = {"typeOfExamination", "reasonCode", "treatmentResultId", "dischargeStatusId"}
# API targets for visual acuity (0–10 integer, soft warning)
_EYE_TARGETS = {"leftEyeGlasses", "leftEyeNoGlasses", "rightEyeGlasses", "rightEyeNoGlasses"}
# API targets for body circumferences (20–250 cm, soft warning)
_CIRC_TARGETS = {"waistCircumference", "chestCircumference"}

_TYPE_HINT = {
    "str": "Chữ",
    "int": "Số nguyên",
    "float": "Số thập phân",
    "str_num": "Số",
    "datetime": "Ngày dd/mm/yyyy",
    "list": "Danh sách (mỗi mục một dòng, hoặc phân tách bằng ;)",
}

# Per-field Vietnamese hints, keyed by API target.
_FIELD_HINT = {
    "medicalIdentifierCode": "Mã tra cứu bệnh nhân: CCCD / số thẻ BHYT / mã định danh "
    "(giống ô tìm kiếm trên web).",
    "examinationDate": "Ngày khám. Bỏ trống giờ sẽ dùng giờ mặc định trong cấu hình.",
    "finishExaminationDate": "Ngày/giờ kết thúc khám. Phải sau 'Ngày khám'.",
    "typeOfExamination": "Mã loại khám. Mặc định: 100 (khám ngoại trú).",
    "reasonCode": "Mã lý do khám. Mặc định: 93.",
    "reasonsMedicalexamination": "Lý do khám, ví dụ 'Khám sức khoẻ'.",
    "symptoms": "Triệu chứng, ví dụ 'Không'.",
    "bodySkinDesc": "Để trống = 'Bình thường'.",
    "bodyOtherDesc": "Để trống = 'Bình thường'.",
    "heartDesc": "Để trống = 'Bình thường'.",
    "respiratoryDesc": "Để trống = 'Bình thường'.",
    "digestDesc": "Để trống = 'Bình thường'.",
    "urologyDesc": "Để trống = 'Bình thường'.",
    "nerveDesc": "Để trống = 'Bình thường'.",
    "osteoarthritisDesc": "Để trống = 'Bình thường'.",
    "endocrineDesc": "Để trống = 'Bình thường'.",
    "bloodDesc": "Để trống = 'Bình thường'.",
    "surgeryDesc": "Để trống = 'Bình thường'.",
    "maternityDesc": "Để trống = 'Bình thường'. Điền N/A nếu không áp dụng.",
    "earNoseThroatDesc": "Để trống = 'Bình thường'.",
    "toothDesc": "Để trống = 'Bình thường'.",
    "eyeDesc": "Mô tả khám mắt (phân biệt với thị lực số). Để trống = 'Bình thường'.",
    "dermatologyDesc": "Để trống = 'Bình thường'.",
    "nutritionDesc": "Để trống = 'Bình thường'.",
    "motionDesc": "Để trống = 'Bình thường'.",
    "physicalDesc": "Để trống = 'Bình thường'.",
    "organsOtherDesc": "Để trống = 'Bình thường'.",
    "diagnosesDischarge": "Chẩn đoán ra viện, ví dụ '0000 - Bình thường'.",
    "diagnosesDischargeList": "Danh sách chẩn đoán. Để trống sẽ tự sao chép từ cột 'Chẩn đoán'. "
    "Nhiều mục: mỗi mục một dòng hoặc phân tách bằng ;",
    "noteDisease": "Ghi chú bệnh lý, ví dụ 'Không'.",
    "treatmentDirection": "Hướng điều trị. Để trống = 'Bình thường'.",
    "treatmentResultId": "Mã kết quả điều trị (số nguyên). Mặc định: 3.",
    "dischargeStatusId": "Mã tình trạng ra viện (số nguyên). Mặc định: 1.",
    "doctorName": "Tên bác sĩ. Để trống dùng giá trị mặc định trong cấu hình.",
    "pulse": "Mạch (lần/phút), ví dụ 80.",
    "temperature": "Nhiệt độ °C, ví dụ 36.8.",
    "bloodPressureMax": "Huyết áp tối đa (tâm thu), ví dụ 110.",
    "bloodPressureMin": "Huyết áp tối thiểu (tâm trương), ví dụ 70.",
    "breath": "Nhịp thở (lần/phút), ví dụ 20.",
    "weight": "Cân nặng (kg).",
    "height": "Chiều cao (cm).",
    "bmi": "ĐỂ TRỐNG để tự tính từ cân nặng & chiều cao.",
    "waistCircumference": "Vòng bụng (cm).",
    "chestCircumference": "Vòng ngực (cm).",
    "leftEyeGlasses": "Thị lực mắt trái (có kính).",
    "leftEyeNoGlasses": "Thị lực mắt trái (không kính).",
    "rightEyeGlasses": "Thị lực mắt phải (có kính).",
    "rightEyeNoGlasses": "Thị lực mắt phải (không kính).",
}

# Example values keyed by API target (used to fill demo rows).
_NORMAL = "Bình thường"
_EXAMPLE: dict[str, list[Any]] = {
    "medicalIdentifierCode": ["027148003240", "2720551044"],
    "examinationDate": [dt.datetime(2026, 6, 17, 7, 0), dt.datetime(2026, 6, 17, 7, 30)],
    "finishExaminationDate": [dt.datetime(2026, 6, 17, 8, 30), dt.datetime(2026, 6, 17, 9, 0)],
    "typeOfExamination": [100, 100],
    "reasonCode": [93, 93],
    "reasonsMedicalexamination": ["Khám sức khoẻ", "Khám sức khoẻ"],
    "symptoms": ["Không", "Không"],
    "bodySkinDesc": [_NORMAL, _NORMAL],
    "bodyOtherDesc": [_NORMAL, _NORMAL],
    "heartDesc": [_NORMAL, _NORMAL],
    "respiratoryDesc": [_NORMAL, _NORMAL],
    "digestDesc": [_NORMAL, _NORMAL],
    "urologyDesc": [_NORMAL, _NORMAL],
    "nerveDesc": [_NORMAL, _NORMAL],
    "osteoarthritisDesc": [_NORMAL, _NORMAL],
    "endocrineDesc": [_NORMAL, _NORMAL],
    "bloodDesc": [_NORMAL, _NORMAL],
    "surgeryDesc": [_NORMAL, _NORMAL],
    "maternityDesc": [_NORMAL, _NORMAL],
    "earNoseThroatDesc": [_NORMAL, _NORMAL],
    "toothDesc": [_NORMAL, _NORMAL],
    "eyeDesc": [_NORMAL, _NORMAL],
    "dermatologyDesc": [_NORMAL, _NORMAL],
    "nutritionDesc": [_NORMAL, _NORMAL],
    "motionDesc": [_NORMAL, _NORMAL],
    "physicalDesc": [_NORMAL, _NORMAL],
    "organsOtherDesc": [_NORMAL, _NORMAL],
    "diagnosesDischarge": ["0000 - Bình thường", "0000 - Bình thường"],
    "diagnosesDischargeList": ["0000 - Bình thường", "0000 - Bình thường"],
    "noteDisease": ["Không", "Không"],
    "treatmentDirection": [_NORMAL, _NORMAL],
    "treatmentResultId": [3, 3],
    "dischargeStatusId": [1, 1],
    "doctorName": ["Nguyễn Thị Hoa", "Nguyễn Thị Hoa"],
    "pulse": [80, 78],
    "temperature": [36.8, 37.0],
    "bloodPressureMax": [110, 120],
    "bloodPressureMin": [70, 80],
    "breath": [20, 19],
    "weight": [18, 55],
    "height": [140, 160],
    "bmi": [None, None],
    "waistCircumference": [60, 72],
    "chestCircumference": [60, 80],
    "leftEyeGlasses": [10, 10],
    "leftEyeNoGlasses": [10, 8],
    "rightEyeGlasses": [10, 10],
    "rightEyeNoGlasses": [10, 9],
}

# Vitals reference lines for the guide, derived from _RANGES at module load time.
_VITALS_LABEL = {
    "pulse": "Mạch (lần/phút)",
    "temperature": "Nhiệt độ (°C)",
    "bloodPressureMax": "HA tối đa (mmHg)",
    "bloodPressureMin": "HA tối thiểu (mmHg)",
    "breath": "Nhịp thở (lần/phút)",
    "weight": "Cân nặng (kg)",
    "height": "Chiều cao (cm)",
}
_RANGES_LINE = "  ".join(
    f"{_VITALS_LABEL.get(k, k)} {lo}–{hi}" for k, (lo, hi) in _RANGES.items() if k in _VITALS_LABEL
)

_GUIDE = [
    "HƯỚNG DẪN",
    "",
    "• Mỗi dòng = 1 bệnh nhân (1 lần khám sức khoẻ).",
    "• KHÔNG đổi tên các cột ở dòng tiêu đề — ứng dụng dựa vào đó để đọc dữ liệu. "
    "Dòng tiêu đề được khoá (bảo vệ); chọn Review → Unprotect Sheet nếu cần sửa cấu trúc.",
    "• Màu cột tiêu đề: CAM = mã định danh (bắt buộc, dùng tìm bệnh nhân); "
    "ĐỎ = bắt buộc (để trống sẽ bị lỗi); XANH = tuỳ chọn (để trống dùng giá trị mặc định).",
    "• Cột 'Mã định danh': nhập CCCD hoặc số thẻ BHYT để tìm bệnh nhân (giống ô tìm "
    "kiếm trên web). Không cần là mã định danh y tế.",
    "• Ngày: định dạng dd/mm/yyyy, có thể kèm giờ. Để trống giờ sẽ dùng giờ mặc định.",
    "• Số thập phân: có thể dùng dấu phẩy hoặc dấu chấm (36,8 hoặc 36.8 đều hợp lệ).",
    "• Cột 'BMI': để TRỐNG, ứng dụng tự tính từ cân nặng và chiều cao.",
    "• Mẫu hỗ trợ tối đa 1000 dòng dữ liệu có kiểm tra (validation). "
    "Để nhập thêm: chọn Review → Unprotect Sheet.",
    "⚠️ XOÁ các dòng ví dụ (nền vàng nhạt) trước khi chạy thật — nếu không, chúng sẽ bị "
    "đẩy lên hệ thống như bệnh nhân thật.",
    "• Di chuột vào ô tiêu đề để xem chú thích và ví dụ từng cột.",
    "",
    f"• Khoảng tham chiếu sinh hiệu (cảnh báo nếu ngoài khoảng, không chặn): {_RANGES_LINE}.",
    "",
    "Quy trình: Mở app → Đăng nhập → Chọn Excel → Validate → Dry-run (xem trước) → bỏ "
    "Dry-run, Limit=1 để thử 1 bệnh nhân → kiểm tra trên web → chạy toàn bộ.",
]


def _fmt_example(value: object) -> str | None:
    """Format an example value for a header comment; return None to skip."""
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)


def _style_data_block(
    ws: Worksheet, mapping: MappingConfig, headers: list[str], *, protect: bool
) -> None:
    """Apply 12pt font, thin borders, per-column number format, and optional unlock to the block."""
    last_row = 1 + _DATA_ROWS
    font_12 = Font(size=12)
    unlocked = Protection(locked=False)
    id_col = mapping.identifier.column
    formats = [_number_format(mapping.columns[h], is_identifier=(h == id_col)) for h in headers]
    for row in range(2, last_row + 1):
        for c_idx, fmt in enumerate(formats, start=1):
            cell = ws.cell(row=row, column=c_idx)
            cell.font = font_12
            cell.border = _BORDER_CELL
            if fmt is not None:
                cell.number_format = fmt
            if protect:
                cell.protection = unlocked


def _add_validations(
    ws: Worksheet,
    mapping: MappingConfig,
    headers: list[str],
    *,
    exam_date_col: str | None,
) -> None:
    """Attach Excel data-validation rules to the data region (rows 2..1+_DATA_ROWS)."""
    last_row = 1 + _DATA_ROWS

    for c_idx, header in enumerate(headers, start=1):
        spec = mapping.columns[header]
        target = spec.target
        letter = get_column_letter(c_idx)
        cell_range = f"{letter}2:{letter}{last_row}"

        dv: DataValidation | None = None

        if target in _CODE_INT_TARGETS:
            dv = DataValidation(
                type="whole",
                operator="greaterThanOrEqual",
                formula1="0",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Mã không hợp lệ",
                error="Trường này yêu cầu số nguyên không âm (ví dụ: 100, 93, 3, 1).",
                showInputMessage=False,
            )

        elif target in _RANGES and spec.type in ("int",):
            lo, hi = _RANGES[target]
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=str(int(lo)),
                formula2=str(int(hi)),
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                errorTitle=header,
                error=f"Giá trị nên trong khoảng {lo}–{hi}.",
                showInputMessage=False,
            )

        elif target in _RANGES and spec.type in ("float", "str_num"):
            lo, hi = _RANGES[target]
            dv = DataValidation(
                type="decimal",
                operator="between",
                formula1=str(lo),
                formula2=str(hi),
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                errorTitle=header,
                error=f"Giá trị nên trong khoảng {lo}–{hi}.",
                showInputMessage=False,
            )

        elif target in _EYE_TARGETS:
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1="0",
                formula2="10",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                errorTitle=header,
                error="Thị lực thường từ 0 đến 10.",
                showInputMessage=False,
            )

        elif target in _CIRC_TARGETS:
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1="20",
                formula2="250",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                errorTitle=header,
                error="Chu vi thường từ 20 đến 250 cm.",
                showInputMessage=False,
            )

        elif target == "examinationDate":
            dv = DataValidation(
                type="date",
                operator="between",
                formula1="DATE(2000,1,1)",
                formula2="DATE(2100,12,31)",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                errorTitle="Ngày khám",
                error="Vui lòng nhập ngày hợp lệ (dd/mm/yyyy).",
                showInputMessage=False,
            )

        elif target == "finishExaminationDate" and exam_date_col:
            # Cross-field: finishExaminationDate >= examinationDate (mirrors coerce._check_dates)
            dv = DataValidation(
                type="custom",
                formula1=(f'OR(${letter}2="",${exam_date_col}2="",${letter}2>=${exam_date_col}2)'),
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
                errorTitle="Giờ kết thúc",
                error="Giờ kết thúc phải bằng hoặc sau Ngày khám.",
                showInputMessage=False,
            )

        if dv is not None:
            ws.add_data_validation(dv)
            dv.add(cell_range)


def _apply_protection(ws: Worksheet) -> None:
    """Enable sheet-level protection flags (per-cell unlock done in _style_data_block)."""
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.insertColumns = True  # block: would break the column contract
    ws.protection.deleteColumns = True  # block: would break the column contract
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def make_template(
    mapping: MappingConfig,
    out_path: str | Path,
    *,
    examples: bool = True,
    protect: bool = True,
) -> Path:
    """Write an .xlsx template with one column per mapped Excel header, plus a guide sheet."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Dữ liệu"

    headers = list(mapping.columns.keys())
    id_col = mapping.identifier.column

    # Track column letters for cross-field DV
    exam_date_col: str | None = None

    # ── Header row ────────────────────────────────────────────────────────────
    for c, header in enumerate(headers, start=1):
        spec = mapping.columns[header]
        cell = ws.cell(row=1, column=c, value=header)

        if header == id_col:
            cell.fill = _FILL_ID
            cell.font = Font(bold=True, size=12, color=_COLOR_ID)
        elif spec.required:
            cell.fill = _FILL_REQ
            cell.font = Font(bold=True, size=12, color=_COLOR_REQ)
        else:
            cell.fill = _FILL_OPT
            cell.font = Font(bold=True, size=12, color=_COLOR_OPT)

        cell.border = _BORDER_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        note = f"Trường API: {spec.target}\nKiểu: {_TYPE_HINT.get(spec.type, spec.type)}"
        if spec.required or header == id_col:
            note += "\n(Bắt buộc)"
        hint = _FIELD_HINT.get(spec.target)
        if hint:
            note += f"\n{hint}"
        ex_val = _fmt_example((_EXAMPLE.get(spec.target) or [None])[0])
        if ex_val is not None:
            note += f"\nVí dụ: {ex_val}"
        if header == id_col:
            note += "\n\nXem sheet 'Hướng dẫn' (tab màu cam phía dưới) để biết cách điền."
        cell.comment = Comment(note, "hssk-tools")

        if spec.target == "examinationDate":
            exam_date_col = get_column_letter(c)

        # Compute width: max of header length and longest example value
        ex_vals = _EXAMPLE.get(spec.target) or []
        max_ex_len = max(
            (len(_fmt_example(v) or "") for v in ex_vals if v is not None),
            default=0,
        )
        width = max(12, min(40, max(len(header) + 4, max_ex_len + 2)))
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Collapsible organ-description group ───────────────────────────────────
    desc_indices = [
        c for c, h in enumerate(headers, start=1) if mapping.columns[h].target.endswith("Desc")
    ]
    if desc_indices:
        is_contiguous = desc_indices == list(range(desc_indices[0], desc_indices[-1] + 1))
        if is_contiguous:
            first_letter = get_column_letter(desc_indices[0])
            last_letter = get_column_letter(desc_indices[-1])
            ws.column_dimensions.group(  # type: ignore[attr-defined]
                first_letter, last_letter, outline_level=1, hidden=False
            )

    # ── Data block: 12pt font + borders + number formats + optional unlock ───
    _style_data_block(ws, mapping, headers, protect=protect)

    # ── Example rows (override font to italic 12pt; border already set above) ─
    if examples:
        example_font = Font(italic=True, size=12, color="595959")
        example_fill = PatternFill("solid", fgColor="FFF2CC")
        example_note = (
            "DÒNG VÍ DỤ — XOÁ dòng này trước khi chạy thật. Nếu giữ lại, dữ liệu ví dụ "
            "sẽ bị đẩy lên hệ thống như dữ liệu thật."
        )
        n = len(next(iter(_EXAMPLE.values())))
        for i in range(n):
            row_idx = i + 2  # rows 2, 3, ...
            for c, header in enumerate(headers, start=1):
                target = mapping.columns[header].target
                value = (_EXAMPLE.get(target) or [None] * n)[i]
                cell = ws.cell(row=row_idx, column=c, value=value)
                cell.font = example_font
                cell.fill = example_fill
            ws.cell(row=row_idx, column=1).comment = Comment(example_note, "hssk-tools")

    # ── Data validation ───────────────────────────────────────────────────────
    _add_validations(ws, mapping, headers, exam_date_col=exam_date_col)

    # ── Sheet protection ──────────────────────────────────────────────────────
    if protect:
        _apply_protection(ws)

    # ── Guide sheet ───────────────────────────────────────────────────────────
    guide = wb.create_sheet("Hướng dẫn")
    guide.sheet_properties.tabColor = "FFC000"  # amber tab — matches identifier color family
    guide.column_dimensions["A"].width = 110
    guide.column_dimensions["B"].width = 30

    for r, line in enumerate(_GUIDE, start=1):
        cell = guide.cell(row=r, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith("⚠️"):
            cell.font = Font(bold=True, size=12, color=_COLOR_REQ)
        else:
            cell.font = Font(size=12)

    # Color legend swatches (placed after the guide text, with a blank separator)
    legend_start = len(_GUIDE) + 2
    legend_items = [
        (_FILL_ID, _COLOR_ID, "CAM — Mã định danh (bắt buộc, dùng tìm bệnh nhân)"),
        (_FILL_REQ, _COLOR_REQ, "ĐỎ — Bắt buộc (để trống sẽ bị lỗi)"),
        (_FILL_OPT, _COLOR_OPT, "XANH — Tuỳ chọn (để trống dùng giá trị mặc định)"),
    ]
    legend_title = guide.cell(row=legend_start, column=1, value="Chú giải màu tiêu đề:")
    legend_title.font = Font(bold=True, size=12)
    for offset, (fill, text_color, label) in enumerate(legend_items, start=1):
        swatch = guide.cell(row=legend_start + offset, column=1, value="   ")
        swatch.fill = fill
        swatch.alignment = Alignment(horizontal="center")
        lbl_cell = guide.cell(row=legend_start + offset, column=2, value=label)
        lbl_cell.font = Font(size=12, color=text_color, bold=True)

    wb.save(out)
    return out
