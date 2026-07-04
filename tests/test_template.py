from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from hssk.excel import reader
from hssk.excel.coerce import coerce_row
from hssk.excel.template import _FILL_ID, _FILL_OPT, _FILL_REQ, make_template


def _col_for_target(mapping, target: str) -> str:
    """Excel column letter of the (first) header mapped to ``target``."""
    for i, (_, spec) in enumerate(mapping.columns.items(), start=1):
        if spec.target == target:
            return get_column_letter(i)
    raise AssertionError(f"no column mapped to {target}")


def test_template_headers_match_mapping(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx")
    assert out.exists()
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]
    headers = [c.value for c in ws[1]]
    assert headers == list(mapping.columns.keys())
    assert "Hướng dẫn" in wb.sheetnames


def test_template_rows_read_and_coerce_cleanly(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx", examples=True)
    rows = reader.read_rows(out, mapping)
    assert len(rows) >= 1
    for idx, raw in rows:
        result = coerce_row(raw, mapping, idx)
        assert result.ok, result.errors
        # BMI is left blank in the template and auto-calculated.
        assert result.values.get("bmi")


def test_template_without_examples_is_header_only(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx", examples=False)
    rows = reader.read_rows(out, mapping)
    assert rows == []


def test_template_has_data_validations(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx")
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]
    dvs = list(ws.data_validations.dataValidation)
    assert dvs, "expected at least one DataValidation rule"

    types = {dv.type for dv in dvs}
    assert "whole" in types, "expected whole-number DV for int fields"

    # At least one warning-style between DV for vitals
    warning_between = [dv for dv in dvs if dv.errorStyle == "warning" and dv.operator == "between"]
    assert warning_between, "expected a warning/between DV for vital fields"

    # Code-int fields get a stop-style whole-number DV
    stop_whole = [dv for dv in dvs if dv.errorStyle == "stop" and dv.type == "whole"]
    assert stop_whole, "expected a stop/whole DV for code-int fields"

    # temperature gets a decimal between DV
    decimal_dvs = [dv for dv in dvs if dv.type == "decimal"]
    assert decimal_dvs, "expected a decimal DV for temperature/float fields"


def test_template_protection(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx", protect=True)
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]
    assert ws.protection.sheet, "sheet should be protected"
    # Header row stays locked (default Protection(locked=True))
    assert ws["A1"].protection.locked, "header cell A1 should be locked"
    # Data cells are unlocked
    assert not ws["A2"].protection.locked, "data cell A2 should be unlocked"


def test_template_no_protect_flag(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx", protect=False)
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]
    assert not ws.protection.sheet, "sheet should not be protected when protect=False"


def test_template_guide_has_legend(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx")
    wb = load_workbook(out)
    guide = wb["Hướng dẫn"]

    # Collect all fills used in the guide sheet
    fills = {
        cell.fill.fgColor.rgb
        for row in guide.iter_rows()
        for cell in row
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb"
    }

    assert _FILL_ID.fgColor.rgb in fills, "orange (identifier) swatch missing from guide"
    assert _FILL_REQ.fgColor.rgb in fills, "dark-red (required) swatch missing from guide"
    assert _FILL_OPT.fgColor.rgb in fills, "blue (optional) swatch missing from guide"

    # The legend label text should appear somewhere in the guide
    all_text = " ".join(str(cell.value) for row in guide.iter_rows() for cell in row if cell.value)
    assert "Bắt buộc" in all_text
    assert "Tuỳ chọn" in all_text


def test_template_number_formats(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx")
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]

    id_col = _col_for_target(mapping, "medicalIdentifierCode")
    # Identifier is Text across the whole data region (rows 2 and the last row 1001).
    assert ws[f"{id_col}2"].number_format == "@"
    assert ws[f"{id_col}1001"].number_format == "@"

    date_col = _col_for_target(mapping, "examinationDate")
    assert ws[f"{date_col}2"].number_format == "dd/mm/yyyy hh:mm"

    pulse_col = _col_for_target(mapping, "pulse")
    assert ws[f"{pulse_col}2"].number_format == "0"

    # weight is str_num — stays General so its decimal DV accepts numeric entry.
    weight_col = _col_for_target(mapping, "weight")
    assert ws[f"{weight_col}2"].number_format == "General"


def test_template_guide_discoverable(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx")
    wb = load_workbook(out)
    guide = wb["Hướng dẫn"]
    assert guide.sheet_properties.tabColor is not None

    ws = wb["Dữ liệu"]
    id_col = _col_for_target(mapping, "medicalIdentifierCode")
    comment = ws[f"{id_col}1"].comment
    assert comment is not None and "Hướng dẫn" in comment.text


def test_template_example_rows_flagged(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx", examples=True)
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]
    id_col = _col_for_target(mapping, "medicalIdentifierCode")
    # Both example rows (2 and 3) carry a delete-me comment on the identifier cell.
    for row in (2, 3):
        comment = ws[f"{id_col}{row}"].comment
        assert comment is not None and "XOÁ" in comment.text


def test_template_styling(mapping, tmp_path):
    out = make_template(mapping, tmp_path / "tpl.xlsx")
    wb = load_workbook(out)
    ws = wb["Dữ liệu"]

    # Header cell: 12pt bold
    h = ws["A1"]
    assert h.font.size == 12
    assert h.font.bold
    # Header bottom border is medium (heavier separator)
    assert h.border.bottom.style == "medium"

    # Data cell: 12pt, has a thin border
    d = ws["A2"]
    assert d.font.size == 12
    assert d.border.left.style == "thin"
