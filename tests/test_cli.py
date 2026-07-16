"""Tests for cli.py — template and validate commands (no network)."""

from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

import httpx
import pytest
import respx
from openpyxl import Workbook

from hssk import licensing
from hssk.cli import main

# Captured before any test monkeypatches licensing.check_license (see _default_licensed below),
# so license-specific tests can restore the real implementation and drive it directly.
_REAL_CHECK_LICENSE = licensing.check_license

EXAMPLE_MAPPING = Path(__file__).resolve().parents[1] / "config" / "mapping.example.yaml"


@pytest.fixture(autouse=True)
def _default_licensed(monkeypatch: pytest.MonkeyPatch) -> None:
    """Stub a granted license for every CLI test by default.

    Plan 012 added a launch-time license gate to `hssk.cli.main`; without this, every pre-existing
    command test below would fail (no license key is installed in these sandboxed HSSK_DATA_DIR/
    HSSK_CONFIG_DIR runs, regardless of the shipped `polar_organization_id` default). Tests that
    exercise licensing itself restore the real function via `_REAL_CHECK_LICENSE` and drive it
    directly.
    """
    monkeypatch.setattr(
        licensing,
        "check_license",
        lambda **_kw: licensing.LicenseCheck(ok=True, source="cache", display_key="TEST-STUB"),
    )


_HEADERS = [
    "Mã định danh",
    "Ngày khám",
    "Giờ kết thúc",
    "Mã hình thức khám",
    "Mã đối tượng khám",
    "Lý do khám",
    "Bệnh sử",
    "Da niêm mạc",
    "Toàn thân khác",
    "Tim mạch",
    "Hô hấp",
    "Tiêu hoá",
    "Thận, tiết niệu",
    "Tâm thần - Thần kinh",
    "Cơ xương khớp",
    "Nội tiết",
    "Bệnh máu",
    "Ngoại khoa",
    "Sản phụ khoa",
    "Tai mũi họng",
    "Răng hàm mặt",
    "Mắt",
    "Da liễu",
    "Dinh dưỡng",
    "Vận động",
    "Đánh giá phát triển thể chất",
    "Cơ quan khác",
    "Chẩn đoán",
    "Bệnh kèm theo",
    "Bệnh theo dõi",
    "Tư vấn điều trị",
    "Mã kết quả khám",
    "Mã tình trạng ra viện",
    "Bác sĩ",
    "Mạch",
    "Nhiệt độ",
    "HA tối đa",
    "HA tối thiểu",
    "Nhịp thở",
    "Cân nặng",
    "Chiều cao",
    "BMI",
    "Vòng bụng",
    "Vòng ngực",
    "Mắt trái (kính)",
    "Mắt trái (không kính)",
    "Mắt phải (kính)",
    "Mắt phải (không kính)",
]

_VALID_ROW = [
    "2700020596A",
    dt.datetime(2026, 6, 17, 7, 0, 0),
    dt.datetime(2026, 6, 17, 8, 0, 0),
    100,
    93,
    "Khám sức khoẻ",
    "Không",
    *["Bình thường"] * 20,
    "0000 - Bình thường",
    None,
    None,
    None,
    3,
    1,
    "Nguyễn Thị Hoa",
    80,
    36.8,
    110,
    70,
    20,
    18,
    140,
    None,
    60,
    60,
    10,
    10,
    10,
    10,
]


def _make_xlsx(tmp_path: Path, rows: list | None = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for row in rows or [_VALID_ROW]:
        ws.append(row)
    p = tmp_path / "data.xlsx"
    wb.save(p)
    return p


def _copy_mapping(tmp_path: Path) -> Path:
    dst = tmp_path / "mapping.yaml"
    shutil.copy(EXAMPLE_MAPPING, dst)
    return dst


# -- template --------------------------------------------------------------------------


def test_cmd_template_creates_file(tmp_path: Path):
    mapping_path = _copy_mapping(tmp_path)
    out = tmp_path / "out.xlsx"
    assert main(["template", "-m", str(mapping_path), "-o", str(out)]) == 0
    assert out.exists()


def test_cmd_template_no_examples(tmp_path: Path):
    mapping_path = _copy_mapping(tmp_path)
    out = tmp_path / "out.xlsx"
    assert main(["template", "-m", str(mapping_path), "-o", str(out), "--no-examples"]) == 0
    assert out.exists()


def _header_row(path: Path) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return headers


def test_cmd_template_update_adds_record_id_column(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """`template --update` merges the overlay so the generated Excel has the Mã hồ sơ column."""
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))  # overlay seeds here, not real config dir
    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        mapping_path = _copy_mapping(tmp_path)
        create_out = tmp_path / "create.xlsx"
        update_out = tmp_path / "update.xlsx"
        assert main(["template", "-m", str(mapping_path), "-o", str(create_out)]) == 0
        assert main(["template", "-m", str(mapping_path), "-o", str(update_out), "--update"]) == 0
        assert "Mã hồ sơ" not in _header_row(create_out)
        assert "Mã hồ sơ" in _header_row(update_out)
    finally:
        _s.cache_clear()


# -- validate --------------------------------------------------------------------------


def test_cmd_validate_valid_file(tmp_path: Path):
    mapping_path = _copy_mapping(tmp_path)
    xlsx = _make_xlsx(tmp_path)
    assert main(["validate", "-m", str(mapping_path), "-i", str(xlsx)]) == 0


def test_cmd_validate_invalid_row(tmp_path: Path):
    mapping_path = _copy_mapping(tmp_path)
    bad_row: list = [None] + _VALID_ROW[1:]  # blank identifier → INVALID
    xlsx = _make_xlsx(tmp_path, rows=[bad_row])
    assert main(["validate", "-m", str(mapping_path), "-i", str(xlsx)]) == 1


def test_cmd_validate_missing_input(tmp_path: Path):
    mapping_path = _copy_mapping(tmp_path)
    assert main(["validate", "-m", str(mapping_path), "-i", str(tmp_path / "no.xlsx")]) == 1


def test_cmd_validate_warns_on_extra_column(tmp_path: Path, capsys: pytest.CaptureFixture[str]):
    """An unmapped Excel column is reported as a warning without failing validation."""
    mapping_path = _copy_mapping(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append([*_HEADERS, "Cột lạ"])
    ws.append([*_VALID_ROW, "junk"])
    xlsx = tmp_path / "data.xlsx"
    wb.save(xlsx)
    assert main(["validate", "-m", str(mapping_path), "-i", str(xlsx)]) == 0
    out = capsys.readouterr().out
    assert "Cột lạ" in out
    assert "unmapped Excel column" in out


def test_cmd_validate_bad_mapping(tmp_path: Path):
    bad_mapping = tmp_path / "bad.yaml"
    bad_mapping.write_text("not: valid: yaml: mapping\n", encoding="utf-8")
    assert main(["validate", "-m", str(bad_mapping), "-i", str(_make_xlsx(tmp_path))]) == 1


# -- run (dry-run with mocked token and network) ---------------------------------------


BASE = "https://api.test"


@respx.mock
def test_cmd_run_dry_run(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """run without --commit should write payloads and not call the create endpoint."""
    from hssk.api import exams, patients

    monkeypatch.setenv("HSSK_BASE_URL", BASE)
    monkeypatch.setenv("HSSK_REQUEST_DELAY", "0")
    monkeypatch.setenv("HSSK_JITTER", "0")
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()

    respx.post(f"{BASE}{patients.SEARCH_PATH}").mock(
        return_value=httpx.Response(
            200,
            json={"data": {"content": [{"patientId": 1, "medicalIdentifierCode": "2700020596A"}]}},
        )
    )
    create_route = respx.post(f"{BASE}{exams.CREATE_PATH}").mock(
        return_value=httpx.Response(200, json={"data": {"medicalRecordId": 99}})
    )

    mapping_path = _copy_mapping(tmp_path)
    xlsx = _make_xlsx(tmp_path)

    # Patch load_valid_token so no real token file is needed.
    monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

    result = main(["run", "-m", str(mapping_path), "-i", str(xlsx)])
    assert result == 0
    assert create_route.call_count == 0  # dry-run — never calls create

    _s.cache_clear()


@respx.mock
def test_cmd_run_commit(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """run --commit --yes should call the create endpoint."""
    from hssk.api import exams, patients

    monkeypatch.setenv("HSSK_BASE_URL", BASE)
    monkeypatch.setenv("HSSK_REQUEST_DELAY", "0")
    monkeypatch.setenv("HSSK_JITTER", "0")
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()

    respx.post(f"{BASE}{patients.SEARCH_PATH}").mock(
        return_value=httpx.Response(
            200,
            json={"data": {"content": [{"patientId": 1, "medicalIdentifierCode": "2700020596A"}]}},
        )
    )
    respx.post(f"{BASE}{exams.CREATE_PATH}").mock(
        return_value=httpx.Response(200, json={"data": {"medicalRecordId": 99}})
    )

    mapping_path = _copy_mapping(tmp_path)
    xlsx = _make_xlsx(tmp_path)

    monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

    result = main(["run", "-m", str(mapping_path), "-i", str(xlsx), "--commit", "--yes"])
    assert result == 0

    _s.cache_clear()


def test_cmd_run_commit_aborts_on_non_interactive_stdin(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    """run --commit without --yes must abort cleanly on closed stdin, not crash with EOFError."""
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()

    mapping_path = _copy_mapping(tmp_path)
    xlsx = _make_xlsx(tmp_path)
    monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

    def _eof(*_a, **_k):
        raise EOFError

    monkeypatch.setattr("builtins.input", _eof)

    # No network mocks: aborting at the prompt must happen before any request is made.
    result = main(["run", "-m", str(mapping_path), "-i", str(xlsx), "--commit"])
    assert result == 1

    _s.cache_clear()


def test_cmd_run_retry_pending_flag_reaches_runner(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """--retry-pending (Plan 004) must reach runner.run as retry_pending=True."""
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()

    mapping_path = _copy_mapping(tmp_path)
    xlsx = _make_xlsx(tmp_path)
    monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

    captured: dict = {}

    def fake_run(*args, **kwargs):
        captured.update(kwargs)
        from hssk.pipeline.results import RunSummary

        return RunSummary(total=0, counts={}, outcomes=[], run_dir=tmp_path)

    monkeypatch.setattr("hssk.cli.runner.run", fake_run)

    result = main(["run", "-m", str(mapping_path), "-i", str(xlsx), "--retry-pending"])
    assert result == 0
    assert captured.get("retry_pending") is True

    _s.cache_clear()


def test_cmd_run_without_retry_pending_flag_defaults_false(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()

    mapping_path = _copy_mapping(tmp_path)
    xlsx = _make_xlsx(tmp_path)
    monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

    captured: dict = {}

    def fake_run(*args, **kwargs):
        captured.update(kwargs)
        from hssk.pipeline.results import RunSummary

        return RunSummary(total=0, counts={}, outcomes=[], run_dir=tmp_path)

    monkeypatch.setattr("hssk.cli.runner.run", fake_run)

    result = main(["run", "-m", str(mapping_path), "-i", str(xlsx)])
    assert result == 0
    assert captured.get("retry_pending") is False

    _s.cache_clear()


@pytest.mark.parametrize("answer", ["y", "yes", "Y", "YES"])
def test_confirm_production_accepts_y_variants(monkeypatch: pytest.MonkeyPatch, answer: str):
    from hssk.cli import _confirm_production

    monkeypatch.setattr("builtins.input", lambda *_a, **_k: answer)
    assert _confirm_production("create") is True


@pytest.mark.parametrize("answer", ["n", "no", "", "sure", "YESS"])
def test_confirm_production_rejects_anything_else(monkeypatch: pytest.MonkeyPatch, answer: str):
    from hssk.cli import _confirm_production

    monkeypatch.setattr("builtins.input", lambda *_a, **_k: answer)
    assert _confirm_production("create") is False


# -- delete ----------------------------------------------------------------------------


def _make_delete_xlsx(tmp_path: Path, record_id: int = 388261169) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Mã định danh", "Mã hồ sơ"])
    ws.append(["2700020596A", record_id])
    p = tmp_path / "delete.xlsx"
    wb.save(p)
    return p


def test_cmd_template_delete_two_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """`template --delete` generates exactly the identifier + Mã hồ sơ columns."""
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))  # overlay seeds here
    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        mapping_path = _copy_mapping(tmp_path)
        out = tmp_path / "delete_template.xlsx"
        assert main(["template", "-m", str(mapping_path), "-o", str(out), "--delete"]) == 0
        assert _header_row(out) == ["Mã định danh", "Mã hồ sơ"]
    finally:
        _s.cache_clear()


def test_cmd_validate_delete_two_col_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        mapping_path = _copy_mapping(tmp_path)
        xlsx = _make_delete_xlsx(tmp_path)
        assert main(["validate", "-m", str(mapping_path), "-i", str(xlsx), "--delete"]) == 0
    finally:
        _s.cache_clear()


@respx.mock
def test_cmd_delete_dry_run(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """delete without --commit fetches detail but never calls the delete endpoint."""
    from hssk.api import records

    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_BASE_URL", BASE)
    monkeypatch.setenv("HSSK_REQUEST_DELAY", "0")
    monkeypatch.setenv("HSSK_JITTER", "0")
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
            return_value=httpx.Response(200, json={"medicalRecordInfo": {"patientId": 1}})
        )
        delete_route = respx.post(f"{BASE}{records.DELETE_PATH}/388261169").mock(
            return_value=httpx.Response(200, json={})
        )
        mapping_path = _copy_mapping(tmp_path)
        xlsx = _make_delete_xlsx(tmp_path)
        monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

        assert main(["delete", "-m", str(mapping_path), "-i", str(xlsx)]) == 0
        assert delete_route.call_count == 0  # dry-run
    finally:
        _s.cache_clear()


@respx.mock
def test_cmd_delete_commit_yes(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """delete --commit --yes calls the delete endpoint once."""
    from hssk.api import records

    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_BASE_URL", BASE)
    monkeypatch.setenv("HSSK_REQUEST_DELAY", "0")
    monkeypatch.setenv("HSSK_JITTER", "0")
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
            return_value=httpx.Response(200, json={"medicalRecordInfo": {"patientId": 1}})
        )
        delete_route = respx.post(f"{BASE}{records.DELETE_PATH}/388261169").mock(
            return_value=httpx.Response(200, json={})
        )
        mapping_path = _copy_mapping(tmp_path)
        xlsx = _make_delete_xlsx(tmp_path)
        monkeypatch.setattr("hssk.auth.token_store.load_valid_token", lambda **kw: "fake-token")

        assert main(["delete", "-m", str(mapping_path), "-i", str(xlsx), "--commit", "--yes"]) == 0
        assert delete_route.call_count == 1
    finally:
        _s.cache_clear()


# -- license gate + `hssk license` subcommand (Plan 012) --------------------------------

_POLAR_ORG = "11111111-1111-1111-1111-111111111111"
_POLAR_VALIDATE_URL = "https://api.polar.sh" + licensing.VALIDATE_PATH


def test_gated_command_blocked_without_license(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """A gated command (template) exits 1 with the license message and writes nothing when no
    license key is installed (the sandboxed HSSK_DATA_DIR here has none)."""
    monkeypatch.setattr(licensing, "check_license", _REAL_CHECK_LICENSE)
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        mapping_path = _copy_mapping(tmp_path)
        out = tmp_path / "out.xlsx"
        result = main(["template", "-m", str(mapping_path), "-o", str(out)])
        assert result == 1
        assert not out.exists()
    finally:
        _s.cache_clear()


@respx.mock
def test_cmd_license_set_writes_cache_and_unblocks_gate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    """`hssk license --set` against a granted key writes the cache and exits 0, and a
    subsequently-run gated command then goes through."""
    monkeypatch.setattr(licensing, "check_license", _REAL_CHECK_LICENSE)
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_POLAR_ORGANIZATION_ID", _POLAR_ORG)

    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        respx.post(_POLAR_VALIDATE_URL).mock(
            return_value=httpx.Response(
                200,
                json={
                    "status": "granted",
                    "expires_at": None,
                    "display_key": "TEST-****",
                    "customer": {"email": "clinic@example.com"},
                },
            )
        )
        assert main(["license", "--set", "SOME-KEY"]) == 0

        from hssk.config import license_cache_path

        assert license_cache_path().exists()

        mapping_path = _copy_mapping(tmp_path)
        out = tmp_path / "out.xlsx"
        assert main(["template", "-m", str(mapping_path), "-o", str(out)]) == 0
        assert out.exists()
    finally:
        _s.cache_clear()


def test_cmd_license_bare_reports_not_licensed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
):
    monkeypatch.setattr(licensing, "check_license", _REAL_CHECK_LICENSE)
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        result = main(["license"])
        assert result == 1
        out = capsys.readouterr().out
        assert "Not licensed" in out
    finally:
        _s.cache_clear()


def test_cmd_license_itself_bypasses_the_gate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
):
    """`hssk license` (requires_license=False) must run and print its own status, never the
    gate's "requires a license" message, even on an unconfigured/unlicensed build."""
    monkeypatch.setattr(licensing, "check_license", _REAL_CHECK_LICENSE)
    monkeypatch.setenv("HSSK_CONFIG_DIR", str(tmp_path))
    monkeypatch.setenv("HSSK_DATA_DIR", str(tmp_path))

    from hssk.config import settings as _s

    _s.cache_clear()
    try:
        main(["license"])
        out = capsys.readouterr().out
        assert "requires a license" not in out
    finally:
        _s.cache_clear()
