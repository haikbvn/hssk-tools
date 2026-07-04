"""Tests for excel/reader.py — sheet selection, header_row offset, column checks."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest
from openpyxl import Workbook

from hssk.errors import ConfigError
from hssk.excel.reader import read_rows
from hssk.mapping import MappingConfig

# -- helpers ---------------------------------------------------------------------------


def _minimal_mapping(
    headers: list[str], sheet: str | None = None, header_row: int = 1
) -> MappingConfig:
    """Build a MappingConfig with one required identifier column and extra str columns."""
    columns: dict = {
        headers[0]: {"target": "medicalIdentifierCode", "type": "str", "required": True}
    }
    for h in headers[1:]:
        columns[h] = {"target": h.lower().replace(" ", "_"), "type": "str"}
    return MappingConfig.model_validate(
        {
            "sheet": sheet,
            "header_row": header_row,
            "identifier": {"column": headers[0]},
            "columns": columns,
        }
    )


def _write_xlsx(
    tmp_path: Path, headers: list[str], rows: list[list], sheet_name: str = "Sheet"
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    p = tmp_path / "data.xlsx"
    wb.save(p)
    return p


# -- basic read ------------------------------------------------------------------------


def test_read_simple(tmp_path: Path):
    headers = ["Mã định danh", "Name"]
    p = _write_xlsx(tmp_path, headers, [["ID001", "Alice"], ["ID002", "Bob"]])
    mapping = _minimal_mapping(headers)
    result = read_rows(p, mapping)
    assert len(result) == 2
    assert result[0] == (2, {"Mã định danh": "ID001", "Name": "Alice"})
    assert result[1] == (3, {"Mã định danh": "ID002", "Name": "Bob"})


def test_blank_rows_are_skipped(tmp_path: Path):
    headers = ["Mã định danh"]
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    ws.append(["ID001"])
    ws.append([None])  # blank row
    ws.append(["ID002"])
    p = tmp_path / "data.xlsx"
    wb.save(p)
    mapping = _minimal_mapping(headers)
    result = read_rows(p, mapping)
    assert len(result) == 2


# -- sheet selection -------------------------------------------------------------------


def test_first_sheet_selected_when_sheet_is_none(tmp_path: Path):
    headers = ["Mã định danh"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    ws.append(["ID001"])
    wb.create_sheet("Other")
    p = tmp_path / "data.xlsx"
    wb.save(p)
    mapping = _minimal_mapping(headers, sheet=None)
    result = read_rows(p, mapping)
    assert len(result) == 1


def test_named_sheet_selected(tmp_path: Path):
    headers = ["Mã định danh"]
    wb = Workbook()
    wb.active.title = "Summary"
    ws2 = wb.create_sheet("Records")
    ws2.append(headers)
    ws2.append(["ID001"])
    p = tmp_path / "data.xlsx"
    wb.save(p)
    mapping = _minimal_mapping(headers, sheet="Records")
    result = read_rows(p, mapping)
    assert len(result) == 1


def test_missing_sheet_raises(tmp_path: Path):
    headers = ["Mã định danh"]
    p = _write_xlsx(tmp_path, headers, [["ID001"]])
    mapping = _minimal_mapping(headers, sheet="NoSuchSheet")
    with pytest.raises(ConfigError, match="not found"):
        read_rows(p, mapping)


# -- header_row offset -----------------------------------------------------------------


def test_header_row_offset(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["ignored", "title", "row"])  # row 1 — not headers
    ws.append(["Mã định danh", "Name"])  # row 2 — actual headers
    ws.append(["ID001", "Alice"])
    p = tmp_path / "data.xlsx"
    wb.save(p)
    mapping = _minimal_mapping(["Mã định danh", "Name"], header_row=2)
    result = read_rows(p, mapping)
    assert len(result) == 1
    assert result[0][1]["Mã định danh"] == "ID001"


# -- missing columns -------------------------------------------------------------------


def test_missing_required_column_raises(tmp_path: Path):
    # Workbook has "ID" but mapping expects "Mã định danh"
    p = _write_xlsx(tmp_path, ["ID", "Name"], [["ID001", "Alice"]])
    mapping = _minimal_mapping(["Mã định danh"])
    with pytest.raises(ConfigError, match="missing mapped column"):
        read_rows(p, mapping)


# -- missing file ----------------------------------------------------------------------


def test_missing_file_raises(tmp_path: Path):
    mapping = _minimal_mapping(["Mã định danh"])
    with pytest.raises(ConfigError, match="not found"):
        read_rows(tmp_path / "no_such.xlsx", mapping)


# -- duplicate headers -----------------------------------------------------------------


def test_duplicate_mapped_header_raises(tmp_path: Path):
    # Two "Name" columns, both mapped — last-wins would silently drop data, so error out.
    p = _write_xlsx(tmp_path, ["Mã định danh", "Name", "Name"], [["ID001", "A", "B"]])
    mapping = _minimal_mapping(["Mã định danh", "Name"])
    with pytest.raises(ConfigError, match="duplicate"):
        read_rows(p, mapping)


def test_duplicate_unmapped_header_ignored(tmp_path: Path):
    # Duplicated "Extra" is not in the mapping, so it never reaches the row dict — no error.
    p = _write_xlsx(tmp_path, ["Mã định danh", "Extra", "Extra"], [["ID001", "x", "y"]])
    mapping = _minimal_mapping(["Mã định danh"])
    result = read_rows(p, mapping)
    assert len(result) == 1


# -- unmapped-column warnings ----------------------------------------------------------


def test_extra_header_warns(tmp_path: Path):
    p = _write_xlsx(tmp_path, ["Mã định danh", "Extra"], [["ID001", "x"]])
    mapping = _minimal_mapping(["Mã định danh"])
    warnings: list[str] = []
    result = read_rows(p, mapping, on_warning=warnings.append)
    assert len(result) == 1  # return value unchanged
    assert len(warnings) == 1
    assert "'Extra'" in warnings[0]
    assert "unmapped Excel column" in warnings[0]


def test_no_callback_stays_silent(tmp_path: Path):
    # Backward compat: extra columns without an on_warning callback must not error.
    p = _write_xlsx(tmp_path, ["Mã định danh", "Extra"], [["ID001", "x"]])
    mapping = _minimal_mapping(["Mã định danh"])
    result = read_rows(p, mapping)
    assert len(result) == 1


# -- datetime values pass through ------------------------------------------------------


def test_datetime_values_preserved(tmp_path: Path):
    headers = ["Mã định danh", "Ngày khám"]
    d = dt.datetime(2026, 6, 17, 7, 0, 0)
    p = _write_xlsx(tmp_path, headers, [["ID001", d]])
    mapping = _minimal_mapping(headers)
    result = read_rows(p, mapping)
    assert result[0][1]["Ngày khám"] == d
