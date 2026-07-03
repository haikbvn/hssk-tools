from __future__ import annotations

import csv
import json

import openpyxl

from hssk.pipeline.runner import RowOutcome, Status
from hssk.report import new_run_dir, write_report


def _outcome(row_index=1, status=Status.CREATED, **kwargs) -> RowOutcome:
    return RowOutcome(
        row_index=row_index,
        identifier="MIC001",
        status=status,
        **kwargs,
    )


def test_new_run_dir_live_has_no_suffix(tmp_path):
    d = new_run_dir(tmp_path, dry_run=False)
    assert d.exists()
    assert "dryrun" not in d.name


def test_new_run_dir_dry_run_has_suffix(tmp_path):
    d = new_run_dir(tmp_path, dry_run=True)
    assert d.name.endswith("-dryrun")


def test_write_report_creates_all_files(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=False)
    write_report(run_dir, [_outcome()], dry_run=False)

    assert (run_dir / "results.csv").exists()
    assert (run_dir / "results.xlsx").exists()
    assert (run_dir / "events.jsonl").exists()


def test_csv_has_correct_columns_and_row(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=False)
    o = _outcome(patient_id=42, record_id=99, message="ok", warnings=["w1", "w2"])
    write_report(run_dir, [o], dry_run=False)

    with (run_dir / "results.csv").open(encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    r = rows[0]
    assert r["row"] == "1"
    assert r["identifier"] == "MIC001"
    assert r["status"] == "CREATED"
    assert r["patientId"] == "42"
    assert r["recordId"] == "99"
    assert r["warnings"] == "w1; w2"


def test_jsonl_has_one_json_per_row(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=False)
    outcomes = [_outcome(row_index=i, status=Status.CREATED) for i in range(1, 4)]
    write_report(run_dir, outcomes, dry_run=False)

    lines = (run_dir / "events.jsonl").read_text(encoding="utf-8").splitlines()
    assert len(lines) == 3
    for i, line in enumerate(lines, start=1):
        rec = json.loads(line)
        assert rec["row"] == i
        assert rec["status"] == "CREATED"


def test_xlsx_has_header_and_data_rows(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=False)
    write_report(run_dir, [_outcome(record_id=7)], dry_run=False)

    wb = openpyxl.load_workbook(run_dir / "results.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == [
        "row",
        "identifier",
        "status",
        "patientId",
        "recordId",
        "message",
        "warnings",
        "timestamp",
    ]
    data = [c.value for c in ws[2]]
    assert data[0] == 1  # row_index
    assert data[4] == 7  # recordId


def test_warnings_joined_with_semicolon(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=False)
    o = _outcome(warnings=["height missing", "BMI not computed"])
    write_report(run_dir, [o], dry_run=False)

    with (run_dir / "results.csv").open(encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["warnings"] == "height missing; BMI not computed"


def test_timestamp_round_trips_all_formats(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=False)
    o = _outcome(timestamp="2026-07-02T10:30:00")
    write_report(run_dir, [o], dry_run=False)

    with (run_dir / "results.csv").open(encoding="utf-8-sig") as f:
        assert list(csv.DictReader(f))[0]["timestamp"] == "2026-07-02T10:30:00"
    rec = json.loads((run_dir / "events.jsonl").read_text(encoding="utf-8").splitlines()[0])
    assert rec["timestamp"] == "2026-07-02T10:30:00"
    ws = openpyxl.load_workbook(run_dir / "results.xlsx").active
    assert [c.value for c in ws[2]][-1] == "2026-07-02T10:30:00"


def test_timestamp_defaults_to_empty():
    assert _outcome().timestamp == ""


def test_xlsx_many_rows_round_trip(tmp_path):
    # Guards the write_only workbook ordering/contents at scale.
    run_dir = new_run_dir(tmp_path, dry_run=False)
    outcomes = [_outcome(row_index=i, record_id=i) for i in range(1, 201)]
    write_report(run_dir, outcomes, dry_run=False)

    ws = openpyxl.load_workbook(run_dir / "results.xlsx").active
    assert ws.max_row == 201  # header + 200 rows
    assert [c.value for c in ws[2]][0] == 1
    assert [c.value for c in ws[201]][0] == 200


def test_empty_run_writes_header_only(tmp_path):
    run_dir = new_run_dir(tmp_path, dry_run=True)
    write_report(run_dir, [], dry_run=True)

    with (run_dir / "results.csv").open(encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 1  # header only
