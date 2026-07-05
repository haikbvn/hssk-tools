from __future__ import annotations

import json
from pathlib import Path

import httpx
import pytest
import respx
import yaml
from openpyxl import Workbook

from hssk.api import records
from hssk.config import Settings
from hssk.errors import ConfigError
from hssk.mapping import filter_for_delete, load_mapping
from hssk.pipeline import runner
from hssk.pipeline.runner import Status
from test_update import _DETAIL_RESPONSE, _make_xlsx

BASE = "https://api.test"


def _settings(tmp: Path) -> Settings:
    return Settings(base_url=BASE, request_delay=0.0, jitter=0.0, data_dir=tmp)


def _make_delete_mapping(tmp: Path):
    """Base create mapping + the medicalRecordId column, filtered to delete's two columns."""
    repo = Path(__file__).resolve().parents[1]
    base = yaml.safe_load((repo / "config" / "mapping.example.yaml").read_text(encoding="utf-8"))
    base["columns"]["Mã hồ sơ"] = {"target": "medicalRecordId", "type": "int", "required": True}
    p = tmp / "mapping_delete.yaml"
    p.write_text(yaml.dump(base, allow_unicode=True), encoding="utf-8")
    return filter_for_delete(load_mapping(p))


def _make_xlsx_2col(tmp: Path, record_id=388261169) -> Path:
    """A slim delete file: just the identifier + medicalRecordId columns."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Mã định danh", "Mã hồ sơ"])
    ws.append(["2700020596A", record_id])
    p = tmp / "delete_in.xlsx"
    wb.save(p)
    return p


@respx.mock
def test_dry_run_writes_marker_no_delete_call(tmp_path):
    mapping = _make_delete_mapping(tmp_path)
    respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
        return_value=httpx.Response(200, json=_DETAIL_RESPONSE)
    )
    delete_route = respx.post(f"{BASE}{records.DELETE_PATH}/388261169").mock(
        return_value=httpx.Response(200, json={})
    )
    xlsx = _make_xlsx_2col(tmp_path)

    summary = runner.run_delete(
        xlsx, mapping, token="t", dry_run=True, settings=_settings(tmp_path)
    )

    assert summary.counts.get(Status.DRY_RUN_OK) == 1
    assert delete_route.call_count == 0
    marker = summary.run_dir / "payloads" / "row_2.json"
    assert marker.exists()
    assert json.loads(marker.read_text(encoding="utf-8"))["medicalRecordId"] == 388261169
    assert (summary.run_dir / "results.xlsx").exists()


@respx.mock
def test_commit_posts_empty_body_delete(tmp_path):
    mapping = _make_delete_mapping(tmp_path)
    respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
        return_value=httpx.Response(200, json=_DETAIL_RESPONSE)
    )
    captured: list[bytes] = []

    def _capture(request, route):
        captured.append(request.content)
        return httpx.Response(200, json={})

    delete_route = respx.post(f"{BASE}{records.DELETE_PATH}/388261169").mock(side_effect=_capture)
    xlsx = _make_xlsx_2col(tmp_path)

    summary = runner.run_delete(
        xlsx, mapping, token="t", dry_run=False, settings=_settings(tmp_path)
    )

    assert summary.counts.get(Status.DELETED) == 1
    assert delete_route.call_count == 1
    assert captured == [b""]  # empty-body POST (content-length: 0)
    assert summary.outcomes[0].record_id == 388261169
    assert summary.outcomes[0].patient_id == 372954970


@respx.mock
def test_fetch_detail_404_is_failed_row_not_abort(tmp_path):
    """An already-deleted / nonexistent id fails at fetch-detail; the batch continues."""
    mapping = _make_delete_mapping(tmp_path)
    respx.get(f"{BASE}{records.DETAIL_PATH}/111").mock(
        return_value=httpx.Response(404, text="gone")
    )
    respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
        return_value=httpx.Response(200, json=_DETAIL_RESPONSE)
    )
    respx.post(f"{BASE}{records.DELETE_PATH}/388261169").mock(
        return_value=httpx.Response(200, json={})
    )
    wb = Workbook()
    ws = wb.active
    ws.append(["Mã định danh", "Mã hồ sơ"])
    ws.append(["2700020596A", 111])
    ws.append(["2700020596A", 388261169])
    xlsx = tmp_path / "two.xlsx"
    wb.save(xlsx)

    summary = runner.run_delete(
        xlsx, mapping, token="t", dry_run=False, settings=_settings(tmp_path)
    )

    assert not summary.aborted
    assert summary.counts.get(Status.FAILED) == 1
    assert summary.counts.get(Status.DELETED) == 1


@respx.mock
def test_delete_endpoint_4xx_is_failed_row(tmp_path):
    mapping = _make_delete_mapping(tmp_path)
    respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
        return_value=httpx.Response(200, json=_DETAIL_RESPONSE)
    )
    respx.post(f"{BASE}{records.DELETE_PATH}/388261169").mock(
        return_value=httpx.Response(400, text="bad request")
    )
    xlsx = _make_xlsx_2col(tmp_path)

    summary = runner.run_delete(
        xlsx, mapping, token="t", dry_run=False, settings=_settings(tmp_path)
    )

    assert not summary.aborted
    assert summary.counts.get(Status.FAILED) == 1


@respx.mock
def test_auth_expired_on_detail_aborts_batch(tmp_path):
    mapping = _make_delete_mapping(tmp_path)
    respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(return_value=httpx.Response(401))
    xlsx = _make_xlsx_2col(tmp_path)

    summary = runner.run_delete(
        xlsx, mapping, token="t", dry_run=False, settings=_settings(tmp_path)
    )

    assert summary.aborted
    assert summary.counts.get(Status.AUTH_EXPIRED) == 1


def test_blank_record_id_is_invalid(tmp_path):
    mapping = _make_delete_mapping(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.append(["Mã định danh", "Mã hồ sơ"])
    ws.append(["2700020596A", None])
    xlsx = tmp_path / "blank.xlsx"
    wb.save(xlsx)

    summary = runner.run_delete(
        xlsx, mapping, token="t", dry_run=False, settings=_settings(tmp_path)
    )

    assert summary.counts.get(Status.INVALID) == 1


def test_run_delete_without_record_id_column_raises(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    mapping = load_mapping(repo / "config" / "mapping.example.yaml")

    with pytest.raises(ConfigError, match="medicalRecordId"):
        runner.run_delete("nonexistent.xlsx", mapping, token="t", settings=_settings(tmp_path))


@respx.mock
def test_full_update_template_excel_works(tmp_path):
    """A full multi-column update-template file still loads against the slim delete mapping."""
    mapping = _make_delete_mapping(tmp_path)
    respx.get(f"{BASE}{records.DETAIL_PATH}/388261169").mock(
        return_value=httpx.Response(200, json=_DETAIL_RESPONSE)
    )
    logs: list[str] = []
    xlsx = _make_xlsx(tmp_path)  # the 49-column update template

    summary = runner.run_delete(
        xlsx,
        mapping,
        token="t",
        dry_run=True,
        settings=_settings(tmp_path),
        callbacks=runner.Callbacks(on_log=logs.append),
    )

    assert summary.counts.get(Status.DRY_RUN_OK) == 1
    assert any("unmapped Excel column" in m for m in logs)
